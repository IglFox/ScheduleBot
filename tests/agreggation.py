import json
from typing import Dict

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

file_path = "template.xlsx"

Subjects_position = [
    [ # first week
        ["C6", "G6", "K6", "O6", "S6", "W6"],
        ["C10", "G10", "K10", "O10", "S10", "W10"],
        ["C14", "G14", "K14", "O14", "S14", "W14"],
        ["C18", "G18", "K18", "O18", "S18", "W18"],
    ],

    [ # second week
        ["C7", "G7", "K7", "O7", "S7", "W7"],
        ["C11", "G11", "K11", "O11", "S11", "W11"],
        ["C15", "G15", "K15", "O15", "S15", "W15"],
        ["C19", "G19", "K19", "O19", "S19", "W19"],
    ],

    [ # third week
        ["C8", "G8", "K8", "O8", "S8", "W8"],
        ["C12", "G12", "K12", "O12", "S12", "W12"],
        ["C16", "G16", "K16", "O16", "S16", "W16"],
        ["C20", "G20", "K20", "O20", "S20", "W20"],
    ],
]

dates_position = [
    ["C5", "G5", "K5", "O5", "S5", "W5"],
    ["C9", "G9", "K9", "O9", "S9", "W9"],
    ["C13", "G13", "K13", "O13", "S13", "W13"],
    ["C17", "G17", "K17", "O17", "S17", "W17"],
]


def offset_cell(cell: str, row_offset: int = 0, col_offset: int = 0) -> str:
    """
    Возвращает новую ячейку, смещённую относительно исходной.
    Пример: offset_cell("C6", 1, 0) → "C7"
    """
    col_str = ''.join(filter(str.isalpha, cell))
    row_str = ''.join(filter(str.isdigit, cell))

    if not col_str or not row_str:
        raise ValueError(f"Некорректный формат ячейки: {cell}")

    col_idx = column_index_from_string(col_str)
    new_col = get_column_letter(col_idx + col_offset)
    new_row = int(row_str) + row_offset

    return f"{new_col}{new_row}"

def get_day_position_by_name(day: str) -> str:
    match day:
        case "Понедельник":
            return Subjects_position[0][0]
        case "Вторник":
            return Subjects_position[0][1]
        case "Среда":
            return Subjects_position[0][2]
        case "Четверг":
            return Subjects_position[0][3]
        case "Пятница":
            return Subjects_position[0][4]
        case "Суббота":
            return Subjects_position[0][5]
        case _:
            raise Exception("Неверный день недели")

def get_next_day_position_by_name(day: str) -> str:
    match day:
        case "Понедельник":
            return Subjects_position[0][1]
        case "Вторник":
            return Subjects_position[0][2]
        case "Среда":
            return Subjects_position[0][3]
        case "Четверг":
            return Subjects_position[0][4]
        case "Пятница":
            return Subjects_position[0][5]
        case "Суббота":
            return Subjects_position[0][1]
        case _:
            raise Exception("Неверный день недели")

def insert_cell(cell: str, value: str) -> None:
    try:
        workbook = load_workbook(file_path)
        sheet = workbook["Лист1"]

        # Записываем значение
        sheet[cell] = value

        # Сохраняем
        workbook.save(file_path)
        print(f"✅ Записано в {cell}: {value}")
    except Exception as e:
        print(f"❌ Ошибка при записи в Excel: {e}")


def insert_data(df: pd.DataFrame, dictionary: dict) -> None:

    for day in df_dict:
        current_day = day["день недели"]
        position = get_day_position_by_name(current_day)  # "С6"
        for i in range(len(current_day['предметы'])):
            position = get_day_position_by_name(current_day)  # "С6"








# Чтение Excel-файла
df_raw = pd.read_excel("schedule_15-05_06-02-2026.xlsx")

pd.set_option('display.max_columns', None)  # Показывать все столбцы
pd.set_option('display.expand_frame_repr', False)  # Не переносить вывод на новую строку
pd.set_option('display.max_colwidth', None)  # Не обрезать содержимое ячеек


df_first_clean = df_raw.drop(columns=['Группы', 'Преподаватель'])
df_dict = df_first_clean.to_dict('records')

df_result: Dict[str, Dict[str, dict]] = {}

for line in df_dict:
    subject = line["Предмет"]
    date_subject = line["Дата"]

    if date_subject not in df_result:
        df_result[date_subject] = {
            "День Недели": line["День недели"],
            "Предметы": {}
        }

    df_result[date_subject]["День Недели"]= line["День недели"]

    df_result[date_subject]['Предметы'][subject] = {
        "Время": line["Время"],
        "Аудитория": line["Аудитория"],
        "Тип": line["ТипЗанятий"]
    }

with open("result_dict.json", "w", encoding='utf-8') as f:
    json.dump(df_result, f, ensure_ascii=False, indent=4)

