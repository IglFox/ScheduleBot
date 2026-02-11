import json
import os
from datetime import datetime
from typing import Dict

import shutil
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from app import config
from app.utils.LoggerHelp import logger_load

logger = logger_load(__name__)

Subjects_Time_positions = [
    [ # first subject
        ["C6", "G6", "K6", "O6", "S6", "W6"],
        ["C10", "G10", "K10", "O10", "S10", "W10"],
        ["C14", "G14", "K14", "O14", "S14", "W14"],
        ["C18", "G18", "K18", "O18", "S18", "W18"],
        ["C22", "G22", "K22", "O22", "S22", "W22"],
    ],

    [ # second subject
        ["C7", "G7", "K7", "O7", "S7", "W7"],
        ["C11", "G11", "K11", "O11", "S11", "W11"],
        ["C15", "G15", "K15", "O15", "S15", "W15"],
        ["C19", "G19", "K19", "O19", "S19", "W19"],
        ["C23", "G23", "K23", "O23", "S23", "W23"],
    ],

    [ # third subject
        ["C8", "G8", "K8", "O8", "S8", "W8"],
        ["C12", "G12", "K12", "O12", "S12", "W12"],
        ["C16", "G16", "K16", "O16", "S16", "W16"],
        ["C20", "G20", "K20", "O20", "S20", "W20"],
        ["C24", "G24", "K24", "O24", "S24", "W24"],
    ],
]

current_week = {
    "понедельник": 0,
    "вторник": 0,
    "среда": 0,
    "четверг": 0,
    "пятница": 0,
    "суббота": 0
}

dayOfWeek_enum = {
    "понедельник": 0,
    "вторник": 1,
    "среда": 2,
    "четверг": 3,
    "пятница": 4,
    "суббота": 5
}

dates_position = [
    ["C5", "G5", "K5", "O5", "S5", "W5"],
    ["C9", "G9", "K9", "O9", "S9", "W9"],
    ["C13", "G13", "K13", "O13", "S13", "W13"],
    ["C17", "G17", "K17", "O17", "S17", "W17"],
    ["C21", "G21", "K21", "O21", "S21", "W21"],
]

check_week = 0

months = {
    1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
    5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
    9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
}

TemplatePath = config.PARSER["TEMPLATE_FILE_PATH"]
RawSchedulePath = config.PARSER["RAW_FILE_PATH"]
CleanSchedulePath = config.PARSER["FILE_PATH"]

def create_file():
    if os.path.exists(TemplatePath) and os.path.exists(RawSchedulePath):
        if os.path.exists(CleanSchedulePath):
            os.remove(CleanSchedulePath)
            shutil.copy(TemplatePath, CleanSchedulePath)
        else:
            shutil.copy(TemplatePath, CleanSchedulePath)
        logger.info(f"Файл {CleanSchedulePath} создан")
        return
    logger.error("Файлы не найдены")
    raise FileNotFoundError

def offset_cell(cell: str, row_offset: int = 0, col_offset: int = 0) -> str:
    """
    Возвращает новую ячейку, смещённую относительно исходной.
    Пример: offset_cell("C6", 1, 0) → "C7"
    """
    col_str = ''.join(filter(str.isalpha, cell))
    row_str = ''.join(filter(str.isdigit, cell))

    if not col_str or not row_str:
        raise ValueError(f"Некорректный формат ячейки: {cell}")

    col_idx = column_index_from_string(col_str)
    new_col = get_column_letter(col_idx + col_offset)
    new_row = int(row_str) + row_offset

    return f"{new_col}{new_row}"

def insert_cell(cell: str, value: str) -> None:
    try:
        workbook = load_workbook(CleanSchedulePath)
        sheet = workbook["Лист1"]

        # Записываем значение
        sheet[cell] = value

        # Сохраняем
        workbook.save(CleanSchedulePath)
        logger.debug(f"✅ Записано в {cell}: {value}")
    except Exception as e:
        logger.error(f"❌ Ошибка при записи в Excel: {e}")

def is_date_in_month(date_str: str, target_month: int) -> bool:
    """Проверяет, находится ли дата в указанном месяце"""
    date_obj = datetime.strptime(date_str, "%d.%m.%Y")
    return date_obj.month == target_month

def count_days_between(start_date: str, end_date: str) -> int:
    if start_date == end_date:
        return 0
    else:
        start = datetime.strptime(start_date, "%d.%m.%Y")
        end = datetime.strptime(end_date, "%d.%m.%Y")
        logger.info(f"Расстояние между {start_date} и {end_date} дней: {(end - start).days}")
        return (end - start).days

def insert_data(dictionary: Dict[str, Dict[str, dict]], month: int) -> None:
    start_date = None

    for dayDate in dictionary.keys():
        if is_date_in_month(dayDate, month+1):
            logger.info("Все даты за месяц записаны")
            break
        elif not is_date_in_month(dayDate, month):
            logger.info(f"Пропускаем {dayDate}")
            continue

        if start_date is not None and dictionary[dayDate]["День Недели"] != "суббота" and count_days_between(start_date, dayDate) >= 2:
            maximum_week = max(current_week.values())

            for day in current_week.keys():
                if current_week[day] < maximum_week:
                    logger.info(f"была {current_week[day]} неделя - стала {maximum_week}")
                    current_week[day] = maximum_week
            start_date = dayDate

        start_date = dayDate

        day = dictionary[dayDate]
        dayOfWeek = day["День Недели"]
        dayOfWeekNum = dayOfWeek_enum[dayOfWeek]
        weekCount = current_week[dayOfWeek]
        current_week[dayOfWeek] += 1
        subjectCount = -1
        dayDatePosition = dates_position[weekCount][dayOfWeekNum]
        insert_cell(dayDatePosition, str(datetime.strptime(dayDate, "%d.%m.%Y").day))

        month_num = datetime.strptime(dayDate, "%d.%m.%Y").month
        month_name = months[month_num]
        insert_cell("C3", month_name)


        for subjectName in day["Предметы"].keys():
            subjectCount += 1
            subject = day["Предметы"][subjectName]

            subjectTimePosition = Subjects_Time_positions[subjectCount][weekCount][dayOfWeekNum]
            subjectTypePosition = offset_cell(subjectTimePosition, 0, 1)
            subjectNamePosition = offset_cell(subjectTimePosition, 0, 2)
            subjectCabinetPosition = offset_cell(subjectTimePosition, 0, 3)

            insert_cell(subjectTimePosition, subject["Время"])
            insert_cell(subjectTypePosition, subject["Тип"])
            insert_cell(subjectNamePosition, subjectName)
            insert_cell(subjectCabinetPosition, subject["Аудитория"])

def sort_subjects_by_time(data):
    """Сортирует предметы по времени внутри каждого дня"""
    for date, day_info in data.items():
        if "Предметы" in day_info and day_info["Предметы"]:
            # Сортируем предметы по времени начала
            sorted_subjects = dict(sorted(
                day_info["Предметы"].items(),
                key=lambda x: x[1]["Время"].split('-')[0]
            ))
            day_info["Предметы"] = sorted_subjects
    return data

def do_dict():
    df_raw = pd.read_excel(RawSchedulePath)

    pd.set_option('display.max_columns', None)  # Показывать все столбцы
    pd.set_option('display.expand_frame_repr', False)  # Не переносить вывод на новую строку
    pd.set_option('display.max_colwidth', None)  # Не обрезать содержимое ячеек


    df_first_clean = df_raw.drop(columns=['Группы', 'Преподаватель'])
    df_dict = df_first_clean.to_dict('records')

    df_result: Dict[str, Dict[str, dict]] = {}

    for line in df_dict:
        subject = line["Предмет"]
        date_subject = line["Дата"]

        if date_subject not in df_result:
            df_result[date_subject] = {
                "День Недели": line["День недели"],
                "Предметы": {}
            }

        df_result[date_subject]["День Недели"]= line["День недели"]

        df_result[date_subject]['Предметы'][subject] = {
            "Время": line["Время"],
            "Аудитория": line["Аудитория"],
            "Тип": line["ТипЗанятий"]
        }
    df_result_sort = sort_subjects_by_time(df_result)

    # with open("result_dict.json", "w", encoding='utf-8') as f:
    #     json.dump(df_result_sort, f, ensure_ascii=False, indent=4)

    return df_result_sort

def clean(month: int):
    create_file()
    dict = do_dict()
    insert_data(dict, month)
    check_week = 0
    for day in current_week.keys():
        current_week[day] = 0
    os.remove(RawSchedulePath)


